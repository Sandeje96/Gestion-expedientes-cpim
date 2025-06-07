import openpyxl
import openpyxl.utils
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
from pathlib import Path
import calendar
from config import BASE_PATH


class TasasAnalyzer:
    def __init__(self, data_manager):
        self.data_manager = data_manager
        
        # Configuración de ingenieros
        self.INGENIEROS = {
            "IMLAUER FERNANDO": ["visado_gas", "visado_salubridad"],
            "ONETTO JOSE": ["visado_electrica", "visado_electromecanica"]
        }
        
        # Porcentajes de distribución
        self.PORCENTAJE_CONSEJO = 0.30
        self.PORCENTAJE_INGENIERO = 0.70
    
    def get_obras_with_visados(self, fecha_inicio=None, fecha_fin=None, incluir_analizadas=False, solo_pagadas=False):
        """Obtiene todas las obras que tienen tasas de visado en el período especificado"""
        try:
            obras = self.data_manager.get_all_works("obra")
            obras_con_visados = []
            
            for obra in obras:
                obra_completa = self.data_manager.get_work_by_id("obra", obra["id"])
                if not obra_completa:
                    continue
                
                # Verificar si tiene alguna tasa de visado
                tiene_visados = any([
                    obra_completa.get("visado_gas") and str(obra_completa["visado_gas"]).strip(),
                    obra_completa.get("visado_salubridad") and str(obra_completa["visado_salubridad"]).strip(),
                    obra_completa.get("visado_electrica") and str(obra_completa["visado_electrica"]).strip(),
                    obra_completa.get("visado_electromecanica") and str(obra_completa["visado_electromecanica"]).strip()
                ])
                
                if not tiene_visados:
                    continue
                
                # Verificar estado de pago
                estado_pago = obra_completa.get("estado_pago_visado", "No pagado")
                es_pagado = estado_pago == "Pagado"
                
                # Si solo queremos pagadas y no está pagada, saltar
                if solo_pagadas and not es_pagado:
                    continue
                
                # LÓGICA CORREGIDA PARA FECHAS:
                # - Para obras PAGADAS: filtrar por fecha de SALIDA (cuando se pagó)
                # - Para obras NO PAGADAS: incluir TODAS (sin filtro de fecha)
                
                if es_pagado and fecha_inicio and fecha_fin:
                    # Para obras pagadas, usar fecha de salida
                    fecha_salida = obra_completa.get("fecha_salida")
                    if fecha_salida:
                        try:
                            fecha_salida_obj = datetime.strptime(fecha_salida, "%d/%m/%Y")
                            if not (fecha_inicio <= fecha_salida_obj <= fecha_fin):
                                continue  # No está en el período de salida
                        except (ValueError, TypeError):
                            continue  # Fecha de salida inválida
                    else:
                        continue  # No tiene fecha de salida pero está marcada como pagada
                
                # Para obras NO pagadas: incluir todas (sin filtro de fecha)
                # porque necesitamos ver el panorama completo de pendientes
                
                # Verificar si ya fue analizada (si no queremos incluir analizadas)
                if not incluir_analizadas:
                    if obra_completa.get("analizada_en_periodo"):
                        continue
                
                obras_con_visados.append(obra_completa)
            
            return obras_con_visados
        except Exception as e:
            print(f"Error al obtener obras con visados: {e}")
            return []
    
    def calcular_totales_por_tipo(self, obras, solo_pagadas=False):
        """Calcula los totales por tipo de visado"""
        totales = {
            "gas": {"total": 0, "pagado": 0, "no_pagado": 0},
            "salubridad": {"total": 0, "pagado": 0, "no_pagado": 0},
            "electrica": {"total": 0, "pagado": 0, "no_pagado": 0},
            "electromecanica": {"total": 0, "pagado": 0, "no_pagado": 0}
        }
        
        for obra in obras:
            estado_pago = obra.get("estado_pago_visado", "No pagado")
            es_pagado = estado_pago == "Pagado"
            
            # Si solo queremos pagadas y no está pagada, saltar
            if solo_pagadas and not es_pagado:
                continue
            
            # Procesar cada tipo de visado
            visados = {
                "gas": obra.get("visado_gas", ""),
                "salubridad": obra.get("visado_salubridad", ""),
                "electrica": obra.get("visado_electrica", ""),
                "electromecanica": obra.get("visado_electromecanica", "")
            }
            
            for tipo, valor in visados.items():
                if valor and str(valor).strip():
                    try:
                        monto = float(valor)
                        totales[tipo]["total"] += monto
                        
                        if es_pagado:
                            totales[tipo]["pagado"] += monto
                        else:
                            totales[tipo]["no_pagado"] += monto
                    except (ValueError, TypeError):
                        continue
        
        return totales
    
    def calcular_por_ingeniero(self, totales, solo_pagadas=True):
        """Calcula los totales por ingeniero"""
        resultado = {}
        
        for ingeniero, tipos_visado in self.INGENIEROS.items():
            total_ingeniero = 0
            
            for tipo in tipos_visado:
                # Convertir nombres de campos
                campo_map = {
                    "visado_gas": "gas",
                    "visado_salubridad": "salubridad", 
                    "visado_electrica": "electrica",
                    "visado_electromecanica": "electromecanica"
                }
                
                campo = campo_map.get(tipo, tipo)
                if campo in totales:
                    if solo_pagadas:
                        total_ingeniero += totales[campo]["pagado"]
                    else:
                        total_ingeniero += totales[campo]["total"]
            
            # Calcular distribución
            consejo = total_ingeniero * self.PORCENTAJE_CONSEJO
            ingeniero_pago = total_ingeniero * self.PORCENTAJE_INGENIERO
            
            resultado[ingeniero] = {
                "total": total_ingeniero,
                "consejo": consejo,
                "ingeniero": ingeniero_pago,
                "tipos": [campo_map.get(t, t) for t in tipos_visado]
            }
        
        return resultado
    
    def generar_analisis_periodo(self, año, mes, marcar_como_analizadas=False):
        """Genera el análisis completo de un período específico"""
        try:
            # Calcular fechas del período
            primer_dia = datetime(año, mes, 1)
            ultimo_dia = datetime(año, mes, calendar.monthrange(año, mes)[1])
            
            # Obtener obras con visados (todas, para luego separar)
            todas_obras_con_visados = self.get_obras_with_visados(incluir_analizadas=False)
            
            if not todas_obras_con_visados:
                return {
                    "error": f"No se encontraron obras con tasas de visado."
                }
            
            # Separar obras pagadas y no pagadas
            obras_pagadas_en_periodo = []
            obras_no_pagadas_todas = []
            
            for obra in todas_obras_con_visados:
                estado_pago = obra.get("estado_pago_visado", "No pagado")
                
                if estado_pago == "Pagado":
                    # Para obras pagadas, verificar si la fecha de salida está en el período
                    fecha_salida = obra.get("fecha_salida")
                    if fecha_salida:
                        try:
                            fecha_salida_obj = datetime.strptime(fecha_salida, "%d/%m/%Y")
                            if primer_dia <= fecha_salida_obj <= ultimo_dia:
                                # Solo agregar si no fue analizada antes
                                if not obra.get("analizada_en_periodo"):
                                    obras_pagadas_en_periodo.append(obra)
                        except (ValueError, TypeError):
                            continue
                else:
                    # Para obras no pagadas, incluir todas
                    obras_no_pagadas_todas.append(obra)
            
            if not obras_pagadas_en_periodo and not obras_no_pagadas_todas:
                return {
                    "error": f"No se encontraron obras pagadas en {calendar.month_name[mes]} {año} que no hayan sido analizadas previamente, ni obras pendientes de pago."
                }
            
            # Calcular totales solo de las obras pagadas en el período
            totales_pagadas = self.calcular_totales_por_tipo(obras_pagadas_en_periodo, solo_pagadas=True)
            
            # Calcular totales generales (incluyendo no pagadas)
            todas_obras = obras_pagadas_en_periodo + obras_no_pagadas_todas
            totales_generales = self.calcular_totales_por_tipo(todas_obras)
            
            # Calcular por ingeniero (solo obras pagadas en el período)
            por_ingeniero = self.calcular_por_ingeniero(totales_pagadas, solo_pagadas=True)
            
            # Marcar obras como analizadas si se solicita (solo las pagadas en el período)
            if marcar_como_analizadas:
                periodo_marca = f"{mes:02d}/{año}"
                for obra in obras_pagadas_en_periodo:
                    self.marcar_obra_como_analizada(obra["id"], periodo_marca)
            
            return {
                "periodo": f"{calendar.month_name[mes]} {año}",
                "fecha_inicio": primer_dia,
                "fecha_fin": ultimo_dia,
                "total_obras": len(todas_obras),
                "obras_pagadas": len(obras_pagadas_en_periodo),
                "obras_no_pagadas": len(obras_no_pagadas_todas),
                "totales_generales": totales_generales,
                "totales_pagadas": totales_pagadas,
                "por_ingeniero": por_ingeniero,
                "obras_detalle": todas_obras,
                "obras_pagadas_detalle": obras_pagadas_en_periodo,
                "obras_no_pagadas_detalle": obras_no_pagadas_todas
            }
            
        except Exception as e:
            print(f"Error al generar análisis: {e}")
            return {"error": f"Error al generar análisis: {str(e)}"}
        
    def generar_analisis_fechas(self, fecha_inicio, fecha_fin, marcar_como_analizadas=False):
        """Genera el análisis completo para un rango de fechas específico"""
        try:
            # Obtener obras con visados (todas, para luego separar)
            todas_obras_con_visados = self.get_obras_with_visados(incluir_analizadas=False)
            
            if not todas_obras_con_visados:
                return {
                    "error": f"No se encontraron obras con tasas de visado."
                }
            
            # Separar obras pagadas y no pagadas
            obras_pagadas_en_periodo = []
            obras_no_pagadas_todas = []
            
            for obra in todas_obras_con_visados:
                estado_pago = obra.get("estado_pago_visado", "No pagado")
                
                if estado_pago == "Pagado":
                    # Para obras pagadas, verificar si la fecha de salida está en el período
                    fecha_salida = obra.get("fecha_salida")
                    if fecha_salida:
                        try:
                            fecha_salida_obj = datetime.strptime(fecha_salida, "%d/%m/%Y")
                            if fecha_inicio <= fecha_salida_obj <= fecha_fin:
                                # Solo agregar si no fue analizada antes
                                if not obra.get("analizada_en_periodo"):
                                    obras_pagadas_en_periodo.append(obra)
                        except (ValueError, TypeError):
                            continue
                else:
                    # Para obras no pagadas, incluir todas
                    obras_no_pagadas_todas.append(obra)
            
            if not obras_pagadas_en_periodo and not obras_no_pagadas_todas:
                fecha_inicio_str = fecha_inicio.strftime("%d/%m/%Y")
                fecha_fin_str = fecha_fin.strftime("%d/%m/%Y")
                return {
                    "error": f"No se encontraron obras pagadas entre {fecha_inicio_str} y {fecha_fin_str} que no hayan sido analizadas previamente, ni obras pendientes de pago."
                }
            
            # Calcular totales solo de las obras pagadas en el período
            totales_pagadas = self.calcular_totales_por_tipo(obras_pagadas_en_periodo, solo_pagadas=True)
            
            # Calcular totales generales (incluyendo no pagadas)
            todas_obras = obras_pagadas_en_periodo + obras_no_pagadas_todas
            totales_generales = self.calcular_totales_por_tipo(todas_obras)
            
            # Calcular por ingeniero (solo obras pagadas en el período)
            por_ingeniero = self.calcular_por_ingeniero(totales_pagadas, solo_pagadas=True)
            
            # Marcar obras como analizadas si se solicita (solo las pagadas en el período)
            if marcar_como_analizadas:
                periodo_marca = f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
                for obra in obras_pagadas_en_periodo:
                    self.marcar_obra_como_analizada(obra["id"], periodo_marca)
            
            # Crear descripción del período
            fecha_inicio_str = fecha_inicio.strftime("%d/%m/%Y")
            fecha_fin_str = fecha_fin.strftime("%d/%m/%Y")
            periodo_descripcion = f"{fecha_inicio_str} - {fecha_fin_str}"
            
            return {
                "periodo": periodo_descripcion,
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin,
                "total_obras": len(todas_obras),
                "obras_pagadas": len(obras_pagadas_en_periodo),
                "obras_no_pagadas": len(obras_no_pagadas_todas),
                "totales_generales": totales_generales,
                "totales_pagadas": totales_pagadas,
                "por_ingeniero": por_ingeniero,
                "obras_detalle": todas_obras,
                "obras_pagadas_detalle": obras_pagadas_en_periodo,
                "obras_no_pagadas_detalle": obras_no_pagadas_todas
            }
            
        except Exception as e:
            print(f"Error al generar análisis por fechas: {e}")
            return {"error": f"Error al generar análisis: {str(e)}"}
    
    def marcar_obra_como_analizada(self, obra_id, periodo):
        """Marca una obra como analizada en un período específico"""
        try:
            data = {"analizada_en_periodo": periodo}
            return self.data_manager.update_obra_general(obra_id, data)
        except Exception as e:
            print(f"Error al marcar obra como analizada: {e}")
            return False
    
    def exportar_a_excel(self, analisis, ruta_archivo):
        """Exporta el análisis a un archivo Excel con auto-ajuste de columnas"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Análisis Tasas Visado"
            
            # Estilos
            titulo_font = Font(bold=True, size=14)
            header_font = Font(bold=True, size=12)
            subtitulo_font = Font(bold=True, size=11)
            normal_font = Font(size=10)
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            subtotal_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            row = 1
            
            # Título del reporte
            sheet.merge_cells(f'A{row}:K{row}')
            cell = sheet[f'A{row}']
            cell.value = f"ANÁLISIS DE TASAS DE VISADO - {analisis['periodo'].upper()}"
            cell.font = titulo_font
            cell.alignment = Alignment(horizontal='center')
            row += 2
            
            # Información general
            sheet[f'A{row}'] = "Período analizado:"
            sheet[f'B{row}'] = analisis['periodo']
            sheet[f'A{row}'].font = header_font
            row += 1
            
            sheet[f'A{row}'] = "Total de obras:"
            sheet[f'B{row}'] = analisis['total_obras']
            row += 1
            
            sheet[f'A{row}'] = "Obras pagadas:"
            sheet[f'B{row}'] = analisis['obras_pagadas']
            row += 1
            
            sheet[f'A{row}'] = "Obras no pagadas:"
            sheet[f'B{row}'] = analisis['obras_no_pagadas']
            row += 2
            
            # Encabezados de la tabla principal
            headers = ["Fecha", "Profesional", "Comitente", "Gas", "Salubridad", "Eléctrica", "Electromecanica", "Total Visados", "Estado Pago", "Fecha Salida", "GOP"]
            
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
            
            # Datos de obras pagadas
            if analisis['obras_pagadas_detalle']:
                # Subtítulo para obras pagadas
                sheet.merge_cells(f'A{row}:K{row}')
                cell = sheet[f'A{row}']
                cell.value = "OBRAS PAGADAS"
                cell.font = subtitulo_font
                cell.fill = subtotal_fill
                row += 1
                
                for obra in analisis['obras_pagadas_detalle']:
                    self._escribir_fila_obra(sheet, row, obra, border)
                    row += 1
            
            # Datos de obras no pagadas
            if analisis['obras_no_pagadas_detalle']:
                row += 1
                # Subtítulo para obras no pagadas
                sheet.merge_cells(f'A{row}:K{row}')
                cell = sheet[f'A{row}']
                cell.value = "OBRAS NO PAGADAS"
                cell.font = subtitulo_font
                cell.fill = subtotal_fill
                row += 1
                
                for obra in analisis['obras_no_pagadas_detalle']:
                    self._escribir_fila_obra(sheet, row, obra, border)
                    row += 1
            
            # Resumen de totales
            row += 2
            self._escribir_resumen_totales(sheet, row, analisis, header_font, header_fill, border)
            
            # APLICAR AUTO-AJUSTE DE COLUMNAS AL FINAL
            self._auto_ajustar_columnas_mejorado(sheet)
            
            # Guardar archivo
            workbook.save(ruta_archivo)
            return str(ruta_archivo)
            
        except Exception as e:
            print(f"Error al exportar a Excel: {e}")
            return None

    def _auto_ajustar_columnas_mejorado(self, sheet):
        """Ajusta automáticamente el ancho de todas las columnas basándose en el contenido"""
        # Obtener el rango de columnas usado
        max_column = sheet.max_column
        
        for col_num in range(1, max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_num)
            
            # Iterar por todas las filas de esta columna
            for row_num in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                
                # Saltar celdas que son parte de un merge (excepto la celda principal)
                if hasattr(cell, 'coordinate'):
                    is_merged = any(
                        cell.coordinate in merged_range 
                        for merged_range in sheet.merged_cells.ranges
                    )
                    if is_merged:
                        # Si es una celda combinada, usar solo la celda principal para calcular
                        merged_range = next(
                            (mr for mr in sheet.merged_cells.ranges if cell.coordinate in mr), 
                            None
                        )
                        if merged_range and cell.coordinate != merged_range.start_cell.coordinate:
                            continue
                
                if cell.value:
                    # Convertir a string y manejar diferentes tipos de contenido
                    if isinstance(cell.value, (int, float)):
                        # Para números con formato de moneda, considerar el texto formateado
                        if cell.number_format and '$' in cell.number_format:
                            # Simular formato de moneda para calcular longitud
                            cell_length = len(f"${cell.value:,.2f}")
                        else:
                            cell_length = len(str(cell.value))
                    else:
                        cell_length = len(str(cell.value))
                    
                    # Ajustar por fuente en negrita (ocupa más espacio)
                    if cell.font and cell.font.bold:
                        cell_length = int(cell_length * 1.2)
                    
                    # Para celdas combinadas, dividir el ancho entre las columnas combinadas
                    if hasattr(cell, 'coordinate'):
                        merged_range = next(
                            (mr for mr in sheet.merged_cells.ranges if cell.coordinate in mr), 
                            None
                        )
                        if merged_range:
                            # Calcular cuántas columnas abarca el merge
                            cols_in_merge = merged_range.max_col - merged_range.min_col + 1
                            if cols_in_merge > 1:
                                cell_length = cell_length // cols_in_merge
                    
                    if cell_length > max_length:
                        max_length = cell_length
            
            # Calcular ancho ajustado con padding adicional
            # Mínimo de 12, máximo de 50 para evitar columnas extremas
            adjusted_width = min(max(max_length + 3, 12), 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
    def _escribir_fila_obra(self, sheet, row, obra, border):
        """Escribe una fila de obra en el Excel"""
        # Calcular total de visados
        gas = float(obra.get("visado_gas", 0) or 0)
        salubridad = float(obra.get("visado_salubridad", 0) or 0)
        electrica = float(obra.get("visado_electrica", 0) or 0)
        electromecanica = float(obra.get("visado_electromecanica", 0) or 0)
        total_visados = gas + salubridad + electrica + electromecanica
        
        datos = [
            obra.get("fecha", ""),
            obra.get("nombre_profesional", ""),
            obra.get("nombre_comitente", ""),
            gas if gas > 0 else "",
            salubridad if salubridad > 0 else "",
            electrica if electrica > 0 else "",
            electromecanica if electromecanica > 0 else "",
            total_visados if total_visados > 0 else "",
            obra.get("estado_pago_visado", ""),
            obra.get("fecha_salida", ""),
            obra.get("nro_sistema_gop", "")
        ]
        
        for col, valor in enumerate(datos, 1):
            cell = sheet.cell(row=row, column=col, value=valor)
            cell.border = border
            
            # Formato para números - FORMATO CORREGIDO
            if col in [4, 5, 6, 7, 8] and isinstance(valor, (int, float)) and valor > 0:
                cell.number_format = '"$"#,##0.00'
    
    def _escribir_resumen_totales(self, sheet, start_row, analisis, header_font, header_fill, border):
        """Escribe el resumen de totales en el Excel con formato mejorado"""
        row = start_row
        
        # Estilos adicionales
        titulo_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        subtitulo_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ingeniero_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        titulo_font = Font(bold=True, size=14, color="FFFFFF")
        subtitulo_font = Font(bold=True, size=12)
        total_font = Font(bold=True, size=11)
        
        # ============ SECCIÓN 2: TOTALES POR TIPO DE VISADO ============
        sheet.merge_cells(f'A{row}:H{row}')
        cell = sheet[f'A{row}']
        cell.value = "💰 TOTALES POR TIPO DE VISADO (Solo obras pagadas en el período)"
        cell.font = titulo_font
        cell.fill = titulo_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        row += 2
        
        # Encabezados de la tabla de tipos
        headers_tipos = ["Tipo de Visado", "Total Pagado", "Ingeniero Responsable"]
        for col, header in enumerate(headers_tipos, 1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = subtitulo_font
            cell.fill = subtitulo_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Datos por tipo de visado
        totales = analisis['totales_pagadas']
        tipos_info = [
            ("Gas", totales["gas"]["pagado"], "IMLAUER FERNANDO"),
            ("Salubridad", totales["salubridad"]["pagado"], "IMLAUER FERNANDO"),
            ("Eléctrica", totales["electrica"]["pagado"], "ONETTO JOSE"),
            ("Electromecanica", totales["electromecanica"]["pagado"], "ONETTO JOSE")
        ]
        
        for tipo, total, ingeniero in tipos_info:
            # Tipo de visado
            cell = sheet.cell(row=row, column=1, value=tipo)
            cell.border = border
            cell.font = Font(size=10)
            
            # Total pagado
            cell = sheet.cell(row=row, column=2, value=total)
            cell.border = border
            cell.number_format = '"$"#,##0.00'
            cell.font = Font(size=10, bold=True)
            
            # Ingeniero responsable
            cell = sheet.cell(row=row, column=3, value=ingeniero)
            cell.border = border
            cell.font = Font(size=10)
            
            row += 1
        
        row += 1
        
        # ============ SECCIÓN 3: CÁLCULO DE HONORARIOS POR INGENIERO ============
        sheet.merge_cells(f'A{row}:H{row}')
        cell = sheet[f'A{row}']
        cell.value = "👷 CÁLCULO DE HONORARIOS POR INGENIERO"
        cell.font = titulo_font
        cell.fill = titulo_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        row += 2
        
        # Encabezados de la tabla de ingenieros
        headers_ingenieros = ["Ingeniero", "Total Tasas", "Para Consejo (30%)", "Para Ingeniero (70%)", "Tipos de Visado"]
        for col, header in enumerate(headers_ingenieros, 1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = subtitulo_font
            cell.fill = subtitulo_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Datos por ingeniero
        por_ingeniero = analisis['por_ingeniero']
        for ingeniero, datos in por_ingeniero.items():
            # Nombre del ingeniero
            cell = sheet.cell(row=row, column=1, value=ingeniero)
            cell.border = border
            cell.fill = ingeniero_fill
            cell.font = Font(size=10, bold=True)
            
            # Total de tasas
            cell = sheet.cell(row=row, column=2, value=datos['total'])
            cell.border = border
            cell.number_format = '"$"#,##0.00'
            cell.font = Font(size=10, bold=True)
            
            # Para el consejo (30%)
            cell = sheet.cell(row=row, column=3, value=datos['consejo'])
            cell.border = border
            cell.number_format = '"$"#,##0.00'
            cell.font = Font(size=10)
            
            # Para el ingeniero (70%)
            cell = sheet.cell(row=row, column=4, value=datos['ingeniero'])
            cell.border = border
            cell.number_format = '"$"#,##0.00'
            cell.font = Font(size=10, bold=True)
            
            # Tipos de visado
            tipos_str = ', '.join([t.title() for t in datos['tipos']])
            cell = sheet.cell(row=row, column=5, value=tipos_str)
            cell.border = border
            cell.font = Font(size=9)
            
            row += 1
        
        row += 1
        
        # ============ SECCIÓN 4: TOTALES GENERALES ============
        sheet.merge_cells(f'A{row}:H{row}')
        cell = sheet[f'A{row}']
        cell.value = "🏛️ TOTALES GENERALES"
        cell.font = titulo_font
        cell.fill = titulo_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        row += 2
        
        # Calcular totales generales
        total_general = sum(d['total'] for d in por_ingeniero.values())
        total_consejo = sum(d['consejo'] for d in por_ingeniero.values())
        total_ingenieros = sum(d['ingeniero'] for d in por_ingeniero.values())
        
        # Tabla de totales generales
        totales_generales_data = [
            ("Total de todas las tasas:", total_general),
            ("Total para el Consejo (30%):", total_consejo),
            ("Total para Ingenieros (70%):", total_ingenieros)
        ]
        
        for label, value in totales_generales_data:
            # Label
            cell = sheet.cell(row=row, column=1, value=label)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            
            # Value
            cell = sheet.cell(row=row, column=2, value=value)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            cell.number_format = '"$"#,##0.00'
            
            row += 1
