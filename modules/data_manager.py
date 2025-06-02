import os
import openpyxl
from openpyxl.styles import Font
from pathlib import Path

class DataManager:
    def __init__(self):
        self.excel_file = Path("registros.xlsx")
        print(f"Usando archivo Excel: {self.excel_file.absolute()}")
        self._ensure_excel_exists()

    def clean_data(self, data):
        """
        Limpia los datos eliminando espacios en blanco innecesarios
        
        Args:
            data: Diccionario con los datos a limpiar
        
        Returns:
            dict: Diccionario con datos limpios
        """
        cleaned_data = {}
        
        for key, value in data.items():
            if isinstance(value, str):
                # Limpiar espacios al inicio y final
                cleaned_value = value.strip()
                cleaned_data[key] = cleaned_value
            else:
                # Si no es string, mantener el valor original
                cleaned_data[key] = value
        
        return cleaned_data
    
    def _ensure_excel_exists(self):
        """Asegura que el archivo Excel existe con las hojas necesarias"""
        try:
            # Primero intenta abrir el archivo existente
            if os.path.exists(self.excel_file):
                try:
                    # Verificar si es un Excel válido
                    workbook = openpyxl.load_workbook(str(self.excel_file))
                    
                    # Verificar que tiene las hojas necesarias
                    necesita_guardar = False
                    
                    if "Obras en general" not in workbook.sheetnames:
                        print("Creando hoja 'Obras en general'")
                        obras_sheet = workbook.create_sheet("Obras en general")
                        headers_obras = [
                            "Fecha", "Profesión", "Formato", "Nro de Copias", "Tipo de trabajo", 
                            "Nombre del Profesional", "Nombre del Comitente", "Ubicación", 
                            "Nro de expte municipal", "Nro de sistema GOP", "Nro de partida inmobiliaria",
                            "Tasa de sellado", "Tasa de visado", 
                            "Visado de instalacion de Gas", "Visado de instalacion de Salubridad", 
                            "Visado de instalacion electrica", "Visado de instalacion electromecanica",
                            "Estado pago sellado", "Estado pago visado",
                            "Nro de expediente CPIM", "Fecha de salida", "Persona que retira", 
                            "Nro de Caja", "Ruta de carpeta", "WhatsApp Profesional", "WhatsApp Tramitador"
                        ]
                        for idx, header in enumerate(headers_obras, 1):
                            cell = obras_sheet.cell(row=1, column=idx, value=header)
                            cell.font = Font(bold=True)
                        necesita_guardar = True
                    else:
                        # Verificar si ya tiene las columnas de WhatsApp
                        obras_sheet = workbook["Obras en general"]
                        if obras_sheet.max_column < 26:  # Si no tiene todas las columnas
                            # Agregar columnas de WhatsApp
                            obras_sheet.cell(row=1, column=25, value="WhatsApp Profesional")
                            obras_sheet.cell(row=1, column=26, value="WhatsApp Tramitador")
                            obras_sheet.cell(row=1, column=25).font = Font(bold=True)
                            obras_sheet.cell(row=1, column=26).font = Font(bold=True)
                            necesita_guardar = True
                    
                    if "Informes técnicos" not in workbook.sheetnames:
                        print("Creando hoja 'Informes técnicos'")
                        informes_sheet = workbook.create_sheet("Informes técnicos")
                        headers_informes = [
                            "Fecha", "Profesión", "Formato", "Nro de Copias", "Tipo de trabajo", 
                            "Detalle", "Profesional", "Comitente", "Tasa de sellado", 
                            "Estado de pago", "Nro de expediente CPIM", "Fecha de salida", 
                            "Persona que retira", "Nro de Caja", "Ruta de carpeta", "WhatsApp Profesional", "WhatsApp Tramitador"
                        ]
                        for idx, header in enumerate(headers_informes, 1):
                            cell = informes_sheet.cell(row=1, column=idx, value=header)
                            cell.font = Font(bold=True)
                        necesita_guardar = True
                    else:
                        # Verificar si ya tiene las columnas de WhatsApp
                        informes_sheet = workbook["Informes técnicos"]
                        if informes_sheet.max_column < 17:  # Si no tiene todas las columnas
                            # Agregar columnas de WhatsApp
                            informes_sheet.cell(row=1, column=16, value="WhatsApp Profesional")
                            informes_sheet.cell(row=1, column=17, value="WhatsApp Tramitador")
                            informes_sheet.cell(row=1, column=16).font = Font(bold=True)
                            informes_sheet.cell(row=1, column=17).font = Font(bold=True)
                            necesita_guardar = True
                    
                    # Guardar si se hicieron cambios
                    if necesita_guardar:
                        workbook.save(str(self.excel_file))
                        print("Archivo Excel actualizado con las hojas necesarias y campos de WhatsApp")
                    else:
                        print("Archivo Excel existente cargado correctamente")
                    
                    return  # El archivo existe y es válido
                except Exception as e:
                    print(f"Error al abrir Excel existente: {e}")
                    # Si hay error, creamos uno nuevo
            
            # Si no existe o hubo error, crear un nuevo archivo Excel
            print("Creando nuevo archivo Excel...")
            workbook = openpyxl.Workbook()
            
            # Configurar la primera hoja
            obras_sheet = workbook.active
            obras_sheet.title = "Obras en general"
            
            # Crear hoja para informes técnicos
            informes_sheet = workbook.create_sheet("Informes técnicos")
            
            # Configurar encabezados para Obras en general
            headers_obras = [
                "Fecha", "Profesión", "Formato", "Nro de Copias", "Tipo de trabajo", 
                "Nombre del Profesional", "Nombre del Comitente", "Ubicación", 
                "Nro de expte municipal", "Nro de sistema GOP", "Nro de partida inmobiliaria",
                "Tasa de sellado", "Tasa de visado", 
                "Visado de instalacion de Gas", "Visado de instalacion de Salubridad", 
                "Visado de instalacion electrica", "Visado de instalacion electromecanica",
                "Estado pago sellado", "Estado pago visado",
                "Nro de expediente CPIM", "Fecha de salida", "Persona que retira", 
                "Nro de Caja", "Ruta de carpeta", "WhatsApp Profesional", "WhatsApp Tramitador"
            ]
            
            # Configurar encabezados para Informes técnicos
            headers_informes = [
                "Fecha", "Profesión", "Formato", "Nro de Copias", "Tipo de trabajo", 
                "Detalle", "Profesional", "Comitente", "Tasa de sellado", 
                "Estado de pago", "Nro de expediente CPIM", "Fecha de salida", 
                "Persona que retira", "Nro de Caja", "Ruta de carpeta", "WhatsApp Profesional", "WhatsApp Tramitador"
            ]
            
            # Aplicar encabezados a Obras en general
            for idx, header in enumerate(headers_obras, 1):
                cell = obras_sheet.cell(row=1, column=idx, value=header)
                cell.font = Font(bold=True)
            
            # Aplicar encabezados a Informes técnicos
            for idx, header in enumerate(headers_informes, 1):
                cell = informes_sheet.cell(row=1, column=idx, value=header)
                cell.font = Font(bold=True)
            
            # Guardar el archivo
            workbook.save(str(self.excel_file))
            print(f"Nuevo archivo Excel creado en: {self.excel_file.absolute()}")
            
        except Exception as e:
            print(f"ERROR CRÍTICO al crear Excel: {e}")
            # Si todo lo demás falla, intentamos con otro nombre
            self.excel_file = Path("datos_cpim.xlsx")
            self._create_basic_excel()
            
    def get_all_profesionales(self):
        """Obtiene la lista de todos los profesionales registrados"""
        try:
            workbook = openpyxl.load_workbook(str(self.excel_file))
            profesionales = set()
            
            # Buscar en la hoja "Obras en general"
            if "Obras en general" in workbook.sheetnames:
                sheet = workbook["Obras en general"]
                for row in range(2, sheet.max_row + 1):
                    profesional = sheet.cell(row=row, column=6).value
                    if profesional:
                        profesionales.add(profesional)
            
            # Buscar en la hoja "Informes técnicos"
            if "Informes técnicos" in workbook.sheetnames:
                sheet = workbook["Informes técnicos"]
                for row in range(2, sheet.max_row + 1):
                    profesional = sheet.cell(row=row, column=7).value
                    if profesional:
                        profesionales.add(profesional)
            
            return sorted(list(profesionales))
        except Exception as e:
            print(f"Error al obtener profesionales: {e}")
            return []

    def get_all_comitentes(self):
        """Obtiene la lista de todos los comitentes registrados"""
        try:
            workbook = openpyxl.load_workbook(str(self.excel_file))
            comitentes = set()
            
            # Buscar en la hoja "Obras en general"
            if "Obras en general" in workbook.sheetnames:
                sheet = workbook["Obras en general"]
                for row in range(2, sheet.max_row + 1):
                    comitente = sheet.cell(row=row, column=7).value
                    if comitente:
                        comitentes.add(comitente)
            
            # Buscar en la hoja "Informes técnicos"
            if "Informes técnicos" in workbook.sheetnames:
                sheet = workbook["Informes técnicos"]
                for row in range(2, sheet.max_row + 1):
                    comitente = sheet.cell(row=row, column=8).value
                    if comitente:
                        comitentes.add(comitente)
            
            return sorted(list(comitentes))
        except Exception as e:
            print(f"Error al obtener comitentes: {e}")
            return []
    
    def _create_basic_excel(self):
        """Crea un Excel básico como último recurso"""
        try:
            wb = openpyxl.Workbook()
            sheet1 = wb.active
            sheet1.title = "Obras en general"
            wb.create_sheet("Informes técnicos")
            wb.save(str(self.excel_file))
            print(f"Excel básico creado en: {self.excel_file}")
        except Exception as e:
            print(f"ERROR FATAL: {e}")
            raise RuntimeError("No se pudo crear el archivo Excel")
    
    def add_obra_general(self, data):
        """Agrega un registro de Obra en general al Excel"""
        try:
            workbook = openpyxl.load_workbook(str(self.excel_file))
            
            # Verificar si la hoja existe, si no, crearla
            if "Obras en general" not in workbook.sheetnames:
                print("Creando hoja 'Obras en general'")
                sheet = workbook.create_sheet("Obras en general")
                # Añadir encabezados
                headers = [
                    "Fecha", "Profesión", "Formato", "Nro de Copias", "Tipo de trabajo", 
                    "Nombre del Profesional", "Nombre del Comitente", "Ubicación", 
                    "Nro de expte municipal", "Nro de sistema GOP", "Nro de partida inmobiliaria",
                    "Tasa de sellado", "Tasa de visado", 
                    "Visado de instalacion de Gas", "Visado de instalacion de Salubridad", 
                    "Visado de instalacion electrica", "Visado de instalacion electromecanica",
                    "Estado pago sellado", "Estado pago visado",
                    "Nro de expediente CPIM", "Fecha de salida", "Persona que retira", 
                    "Nro de Caja", "Ruta de carpeta", "WhatsApp Profesional", "WhatsApp Tramitador"
                ]
                for idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=idx, value=header)
                    cell.font = Font(bold=True)
            else:
                sheet = workbook["Obras en general"]
            
            # Encontrar la próxima fila vacía
            next_row = sheet.max_row + 1
            
            # Mapear los datos a las columnas correspondientes
            columns = {
                1: data.get("fecha", ""),
                2: data.get("profesion", ""),
                3: data.get("formato", ""),
                4: data.get("nro_copias", ""),
                5: data.get("tipo_trabajo", ""),
                6: data.get("nombre_profesional", ""),
                7: data.get("nombre_comitente", ""),
                8: data.get("ubicacion", ""),
                9: data.get("nro_expte_municipal", ""),
                10: data.get("nro_sistema_gop", ""),
                11: data.get("nro_partida_inmobiliaria", ""),
                12: data.get("tasa_sellado", ""),
                13: data.get("tasa_visado", ""),
                14: data.get("visado_gas", ""),
                15: data.get("visado_salubridad", ""),
                16: data.get("visado_electrica", ""),
                17: data.get("visado_electromecanica", ""),
                18: data.get("estado_pago_sellado", "No pagado"),
                19: data.get("estado_pago_visado", "No pagado"),
                20: data.get("nro_expediente_cpim", ""),
                21: data.get("fecha_salida", ""),
                22: data.get("persona_retira", ""),
                23: data.get("nro_caja", ""),
                24: data.get("ruta_carpeta", ""),
                25: data.get("whatsapp_profesional", ""),
                26: data.get("whatsapp_tramitador", "")
            }
            
            # Insertar datos en la hoja
            for col, value in columns.items():
                sheet.cell(row=next_row, column=col, value=value)
            
            workbook.save(str(self.excel_file))
            print(f"Obra agregada en fila {next_row}")
            return next_row - 1  # Retorna el índice del registro (0-based)
        except Exception as e:
            print(f"Error al agregar obra: {e}")
            return -1  # Retorna -1 en caso de error
    
    def add_informe_tecnico(self, data):
        """Agrega un registro de Informe técnico al Excel"""
        try:
            workbook = openpyxl.load_workbook(str(self.excel_file))
            
            # Verificar si la hoja existe, si no, crearla
            if "Informes técnicos" not in workbook.sheetnames:
                print("Creando hoja 'Informes técnicos'")
                sheet = workbook.create_sheet("Informes técnicos")
                # Añadir encabezados
                headers = [
                    "Fecha", "Profesión", "Formato", "Nro de Copias", "Tipo de trabajo", 
                    "Detalle", "Profesional", "Comitente", "Tasa de sellado", 
                    "Estado de pago", "Nro de expediente CPIM", "Fecha de salida", 
                    "Persona que retira", "Nro de Caja", "Ruta de carpeta", "WhatsApp Profesional", "WhatsApp Tramitador"
                ]
                for idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=idx, value=header)
                    cell.font = Font(bold=True)
            else:
                sheet = workbook["Informes técnicos"]
            
            # Encontrar la próxima fila vacía
            next_row = sheet.max_row + 1
            
            # Mapear los datos a las columnas correspondientes
            columns = {
                1: data.get("fecha", ""),
                2: data.get("profesion", ""),
                3: data.get("formato", ""),
                4: data.get("nro_copias", ""),
                5: data.get("tipo_trabajo", ""),
                6: data.get("detalle", ""),
                7: data.get("profesional", ""),
                8: data.get("comitente", ""),
                9: data.get("tasa_sellado", ""),
                10: data.get("estado_pago", "No pagado"),
                11: data.get("nro_expediente_cpim", ""),
                12: data.get("fecha_salida", ""),
                13: data.get("persona_retira", ""),
                14: data.get("nro_caja", ""),
                15: data.get("ruta_carpeta", ""),
                16: data.get("whatsapp_profesional", ""),
                17: data.get("whatsapp_tramitador", "")
            }
            
            # Insertar datos en la hoja
            for col, value in columns.items():
                sheet.cell(row=next_row, column=col, value=value)
            
            workbook.save(str(self.excel_file))
            print(f"Informe agregado en fila {next_row}")
            return next_row - 1  # Retorna el índice del registro (0-based)
        except Exception as e:
            print(f"Error al agregar informe: {e}")
            return -1  # Retorna -1 en caso de error
    
    def get_all_works(self, work_type="obra"):
        """Retorna todos los trabajos del tipo especificado"""
        try:
            workbook = openpyxl.load_workbook(str(self.excel_file))
            
            # Verificar que existan las hojas
            if work_type == "obra":
                hoja = "Obras en general"
            else:
                hoja = "Informes técnicos"
                
            if hoja not in workbook.sheetnames:
                print(f"Hoja '{hoja}' no encontrada, creándola...")
                sheet = workbook.create_sheet(hoja)
                if work_type == "obra":
                    headers = [
                        "Fecha", "Profesión", "Formato", "Nro de Copias", "Tipo de trabajo", 
                        "Nombre del Profesional", "Nombre del Comitente", "Ubicación", 
                        "Nro de expte municipal", "Nro de sistema GOP", "Nro de partida inmobiliaria",
                        "Tasa de sellado", "Tasa de visado", 
                        "Visado de instalacion de Gas", "Visado de instalacion de Salubridad", 
                        "Visado de instalacion electrica", "Visado de instalacion electromecanica",
                        "Estado pago sellado", "Estado pago visado",
                        "Nro de expediente CPIM", "Fecha de salida", "Persona que retira", 
                        "Nro de Caja", "Ruta de carpeta", "WhatsApp Profesional", "WhatsApp Tramitador"
                    ]
                else:
                    headers = [
                        "Fecha", "Profesión", "Formato", "Nro de Copias", "Tipo de trabajo", 
                        "Detalle", "Profesional", "Comitente", "Tasa de sellado", 
                        "Estado de pago", "Nro de expediente CPIM", "Fecha de salida", 
                        "Persona que retira", "Nro de Caja", "Ruta de carpeta", "WhatsApp Profesional", "WhatsApp Tramitador"
                    ]
                
                for idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=idx, value=header)
                    cell.font = Font(bold=True)
                    
                workbook.save(str(self.excel_file))
                return []  # Retornar lista vacía, ya que acabamos de crear la hoja
            
            sheet = workbook[hoja]
            
            works = []
            for row in range(2, sheet.max_row + 1):  # Empezar desde la fila 2 (después de los encabezados)
                work = {}
                
                # Mapear los datos según el tipo de trabajo
                if work_type == "obra":
                    work = {
                        "id": row - 1,
                        "fecha": sheet.cell(row=row, column=1).value,
                        "profesion": sheet.cell(row=row, column=2).value,
                        "nombre_profesional": sheet.cell(row=row, column=6).value,
                        "nombre_comitente": sheet.cell(row=row, column=7).value,
                        "tipo_trabajo": sheet.cell(row=row, column=5).value,
                        "ubicacion": sheet.cell(row=row, column=8).value,
                        "nro_expediente_cpim": sheet.cell(row=row, column=20).value
                    }
                else:
                    work = {
                        "id": row - 1,
                        "fecha": sheet.cell(row=row, column=1).value,
                        "profesion": sheet.cell(row=row, column=2).value,
                        "profesional": sheet.cell(row=row, column=7).value,
                        "comitente": sheet.cell(row=row, column=8).value,
                        "tipo_trabajo": sheet.cell(row=row, column=5).value,
                        "detalle": sheet.cell(row=row, column=6).value,
                        "nro_expediente_cpim": sheet.cell(row=row, column=11).value
                    }
                
                works.append(work)
            
            return works
        except Exception as e:
            print(f"Error al obtener trabajos: {e}")
            return []  # Retorna lista vacía en caso de error
    
    def get_work_by_id(self, work_type, row_id):
        """Obtiene un trabajo específico por ID y tipo"""
        try:
            workbook = openpyxl.load_workbook(str(self.excel_file))
            
            # Verificar que existan las hojas
            if work_type == "obra":
                hoja = "Obras en general"
            else:
                hoja = "Informes técnicos"
                
            if hoja not in workbook.sheetnames:
                print(f"Hoja '{hoja}' no encontrada")
                return None
            
            sheet = workbook[hoja]
            
            # El ID es el índice de fila + 1 para saltar encabezados
            row = row_id + 1
            
            if row > sheet.max_row:
                return None
            
            if work_type == "obra":
                work = {
                    "id": row_id,
                    "fecha": sheet.cell(row=row, column=1).value,
                    "profesion": sheet.cell(row=row, column=2).value,
                    "formato": sheet.cell(row=row, column=3).value,
                    "nro_copias": sheet.cell(row=row, column=4).value,
                    "tipo_trabajo": sheet.cell(row=row, column=5).value,
                    "nombre_profesional": sheet.cell(row=row, column=6).value,
                    "nombre_comitente": sheet.cell(row=row, column=7).value,
                    "ubicacion": sheet.cell(row=row, column=8).value,
                    "nro_expte_municipal": sheet.cell(row=row, column=9).value,
                    "nro_sistema_gop": sheet.cell(row=row, column=10).value,
                    "nro_partida_inmobiliaria": sheet.cell(row=row, column=11).value,
                    "tasa_sellado": sheet.cell(row=row, column=12).value,
                    "tasa_visado": sheet.cell(row=row, column=13).value,
                    "visado_gas": sheet.cell(row=row, column=14).value,
                    "visado_salubridad": sheet.cell(row=row, column=15).value,
                    "visado_electrica": sheet.cell(row=row, column=16).value,
                    "visado_electromecanica": sheet.cell(row=row, column=17).value,
                    "estado_pago_sellado": sheet.cell(row=row, column=18).value,
                    "estado_pago_visado": sheet.cell(row=row, column=19).value,
                    "nro_expediente_cpim": sheet.cell(row=row, column=20).value,
                    "fecha_salida": sheet.cell(row=row, column=21).value,
                    "persona_retira": sheet.cell(row=row, column=22).value,
                    "nro_caja": sheet.cell(row=row, column=23).value,
                    "ruta_carpeta": sheet.cell(row=row, column=24).value,
                    "whatsapp_profesional": sheet.cell(row=row, column=25).value,
                    "whatsapp_tramitador": sheet.cell(row=row, column=26).value
                }
            else:
                work = {
                    "id": row_id,
                    "fecha": sheet.cell(row=row, column=1).value,
                    "profesion": sheet.cell(row=row, column=2).value,
                    "formato": sheet.cell(row=row, column=3).value,
                    "nro_copias": sheet.cell(row=row, column=4).value,
                    "tipo_trabajo": sheet.cell(row=row, column=5).value,
                    "detalle": sheet.cell(row=row, column=6).value,
                    "profesional": sheet.cell(row=row, column=7).value,
                    "comitente": sheet.cell(row=row, column=8).value,
                    "tasa_sellado": sheet.cell(row=row, column=9).value,
                    "estado_pago": sheet.cell(row=row, column=10).value,
                    "nro_expediente_cpim": sheet.cell(row=row, column=11).value,
                    "fecha_salida": sheet.cell(row=row, column=12).value,
                    "persona_retira": sheet.cell(row=row, column=13).value,
                    "nro_caja": sheet.cell(row=row, column=14).value,
                    "ruta_carpeta": sheet.cell(row=row, column=15).value,
                    "whatsapp_profesional": sheet.cell(row=row, column=16).value,
                    "whatsapp_tramitador": sheet.cell(row=row, column=17).value
                }
            
            return work
        except Exception as e:
            print(f"Error al obtener trabajo por ID: {e}")
            return None  # Retorna None en caso de error
    
    def update_obra_general(self, row_id, data):
        """Actualiza un registro de Obra en general en el Excel y opcionalmente actualiza trabajos similares"""
        try:
            workbook = openpyxl.load_workbook(str(self.excel_file))
            
            # Verificar que existe la hoja
            if "Obras en general" not in workbook.sheetnames:
                print("Hoja 'Obras en general' no encontrada")
                return False
            
            sheet = workbook["Obras en general"]
            
            # El ID es el índice de fila + 1 para saltar encabezados
            row = row_id + 1
            
            if row > sheet.max_row:
                return False
            
            # Obtener la obra actual para verificar si hay que actualizar trabajos similares
            obra_actual = self.get_work_by_id("obra", row_id)
            
            # Verificar si se están actualizando campos relacionados con la salida
            campos_salida = ["estado_pago_sellado", "estado_pago_visado", "fecha_salida", "persona_retira", "nro_caja"]
            es_actualizacion_salida = any(campo in data for campo in campos_salida)
            
            # Actualizar solo los campos proporcionados
            for key, value in data.items():
                if key == "tasa_sellado":
                    sheet.cell(row=row, column=12, value=value)
                elif key == "tasa_visado":
                    sheet.cell(row=row, column=13, value=value)
                elif key == "visado_gas":
                    sheet.cell(row=row, column=14, value=value)
                elif key == "visado_salubridad":
                    sheet.cell(row=row, column=15, value=value)
                elif key == "visado_electrica":
                    sheet.cell(row=row, column=16, value=value)
                elif key == "visado_electromecanica":
                    sheet.cell(row=row, column=17, value=value)
                elif key == "estado_pago_sellado":
                    sheet.cell(row=row, column=18, value=value)
                elif key == "estado_pago_visado":
                    sheet.cell(row=row, column=19, value=value)
                elif key == "nro_expediente_cpim":
                    sheet.cell(row=row, column=20, value=value)
                elif key == "fecha_salida":
                    sheet.cell(row=row, column=21, value=value)
                elif key == "persona_retira":
                    sheet.cell(row=row, column=22, value=value)
                elif key == "nro_caja":
                    sheet.cell(row=row, column=23, value=value)
                elif key == "whatsapp_profesional":
                    sheet.cell(row=row, column=25, value=value)
                elif key == "whatsapp_tramitador":
                    sheet.cell(row=row, column=26, value=value)
            
            # Guardar el archivo con los cambios
            workbook.save(str(self.excel_file))
            print(f"Obra actualizada en fila {row}")
            
            # Si se está actualizando datos de salida, buscar y actualizar otros trabajos similares
            if es_actualizacion_salida and obra_actual:
                self._actualizar_trabajos_similares(obra_actual, data)
                
            return True
        except Exception as e:
            print(f"Error al actualizar obra: {e}")
            return False  # Retorna False en caso de error
        
    def _actualizar_trabajos_similares(self, obra_referencia, datos_actualizacion):
        """
        Actualiza otros trabajos que tienen los mismos datos pero diferentes profesionales
        
        Args:
            obra_referencia: Diccionario con los datos de la obra de referencia
            datos_actualizacion: Diccionario con los campos y valores a actualizar
        """
        try:
            # Obtener todas las obras
            todas_obras = self.get_all_works("obra")
            
            # Inicializar contador de obras actualizadas
            obras_actualizadas = 0
            
            for obra in todas_obras:
                # Obtener datos completos
                obra_completa = self.get_work_by_id("obra", obra["id"])
                
                # Verificar si es similar (mismo comitente, ubicación, tipo y no es la misma obra)
                if (obra_completa["id"] != obra_referencia["id"] and
                    obra_completa["nombre_comitente"] == obra_referencia["nombre_comitente"] and
                    obra_completa["ubicacion"] == obra_referencia["ubicacion"] and
                    obra_completa["nro_partida_inmobiliaria"] == obra_referencia["nro_partida_inmobiliaria"]):
                    
                    # Es un trabajo similar, actualizar los campos de salida
                    datos_a_actualizar = {}
                    
                    # Copiar solo los campos de salida que estén en los datos de actualización
                    if "estado_pago_sellado" in datos_actualizacion:
                        datos_a_actualizar["estado_pago_sellado"] = datos_actualizacion["estado_pago_sellado"]
                    
                    if "estado_pago_visado" in datos_actualizacion:
                        datos_a_actualizar["estado_pago_visado"] = datos_actualizacion["estado_pago_visado"]

                    if "nro_expediente_cpim" in datos_actualizacion:
                        datos_a_actualizar["nro_expediente_cpim"] = datos_actualizacion["nro_expediente_cpim"]
                    
                    if "fecha_salida" in datos_actualizacion:
                        datos_a_actualizar["fecha_salida"] = datos_actualizacion["fecha_salida"]
                    
                    if "persona_retira" in datos_actualizacion:
                        datos_a_actualizar["persona_retira"] = datos_actualizacion["persona_retira"]
                    
                    if "nro_caja" in datos_actualizacion:
                        datos_a_actualizar["nro_caja"] = datos_actualizacion["nro_caja"]
                    
                    # Si hay campos para actualizar, hacerlo
                    if datos_a_actualizar:
                        # Evitamos llamar a update_obra_general para no entrar en recursión
                        self._actualizar_obra_sin_recursion(obra_completa["id"], datos_a_actualizar)
                        obras_actualizadas += 1
            
            if obras_actualizadas > 0:
                print(f"Se actualizaron automáticamente {obras_actualizadas} trabajos similares")
            
            return obras_actualizadas
        except Exception as e:
            print(f"Error al actualizar trabajos similares: {e}")
            return 0
        
    def _actualizar_obra_sin_recursion(self, row_id, data):
        """
        Actualiza un registro de Obra en general sin buscar trabajos similares
        (para evitar la recursión)
        """
        try:
            workbook = openpyxl.load_workbook(str(self.excel_file))
            
            # Verificar que existe la hoja
            if "Obras en general" not in workbook.sheetnames:
                return False
            
            sheet = workbook["Obras en general"]
            
            # El ID es el índice de fila + 1 para saltar encabezados
            row = row_id + 1
            
            if row > sheet.max_row:
                return False
            
            # Actualizar solo los campos proporcionados (igual que en update_obra_general)
            for key, value in data.items():
                if key == "tasa_sellado":
                    sheet.cell(row=row, column=12, value=value)
                elif key == "tasa_visado":
                    sheet.cell(row=row, column=13, value=value)
                elif key == "visado_gas":
                    sheet.cell(row=row, column=14, value=value)
                elif key == "visado_salubridad":
                    sheet.cell(row=row, column=15, value=value)
                elif key == "visado_electrica":
                    sheet.cell(row=row, column=16, value=value)
                elif key == "visado_electromecanica":
                    sheet.cell(row=row, column=17, value=value)
                elif key == "estado_pago_sellado":
                    sheet.cell(row=row, column=18, value=value)
                elif key == "estado_pago_visado":
                    sheet.cell(row=row, column=19, value=value)
                elif key == "nro_expediente_cpim":
                    sheet.cell(row=row, column=20, value=value)
                elif key == "fecha_salida":
                    sheet.cell(row=row, column=21, value=value)
                elif key == "persona_retira":
                    sheet.cell(row=row, column=22, value=value)
                elif key == "nro_caja":
                    sheet.cell(row=row, column=23, value=value)
                elif key == "whatsapp_profesional":
                    sheet.cell(row=row, column=25, value=value)
                elif key == "whatsapp_tramitador":
                    sheet.cell(row=row, column=26, value=value)
            
            workbook.save(str(self.excel_file))
            print(f"Trabajo similar actualizado en fila {row}")
            return True
        except Exception as e:
            print(f"Error al actualizar trabajo similar: {e}")
            return False
    
    def update_informe_tecnico(self, row_id, data):
        """Actualiza un registro de Informe técnico en el Excel"""
        try:
            workbook = openpyxl.load_workbook(str(self.excel_file))
            
            # Verificar que existe la hoja
            if "Informes técnicos" not in workbook.sheetnames:
                print("Hoja 'Informes técnicos' no encontrada")
                return False
            
            sheet = workbook["Informes técnicos"]
            
            # El ID es el índice de fila + 1 para saltar encabezados
            row = row_id + 1
            
            if row > sheet.max_row:
                return False
            
            # Actualizar solo los campos proporcionados
            for key, value in data.items():
                if key == "tasa_sellado":
                    sheet.cell(row=row, column=9, value=value)
                elif key == "estado_pago":
                    sheet.cell(row=row, column=10, value=value)
                elif key == "nro_expediente_cpim":
                    sheet.cell(row=row, column=11, value=value)
                elif key == "fecha_salida":
                    sheet.cell(row=row, column=12, value=value)
                elif key == "persona_retira":
                    sheet.cell(row=row, column=13, value=value)
                elif key == "nro_caja":
                    sheet.cell(row=row, column=14, value=value)
                elif key == "whatsapp_profesional":
                    sheet.cell(row=row, column=16, value=value)
                elif key == "whatsapp_tramitador":
                    sheet.cell(row=row, column=17, value=value)
            
            workbook.save(str(self.excel_file))
            print(f"Informe actualizado en fila {row}")
            return True
        except Exception as e:
            print(f"Error al actualizar informe: {e}")
            return False  # Retorna False en caso de error
    
    def get_next_caja_number(self):
        """Obtiene el próximo número de caja disponible"""
        try:
            workbook = openpyxl.load_workbook(str(self.excel_file))
            
            # Revisar en ambas hojas
            max_caja = 0
            
            # Verificar que existan las hojas
            if "Obras en general" in workbook.sheetnames:
                # Revisar en "Obras en general"
                sheet = workbook["Obras en general"]
                for row in range(2, sheet.max_row + 1):
                    caja = sheet.cell(row=row, column=23).value
                    if caja and isinstance(caja, (int, float)):
                        max_caja = max(max_caja, int(caja))
            
            if "Informes técnicos" in workbook.sheetnames:
                # Revisar en "Informes técnicos"
                sheet = workbook["Informes técnicos"]
                for row in range(2, sheet.max_row + 1):
                    caja = sheet.cell(row=row, column=14).value
                    if caja and isinstance(caja, (int, float)):
                        max_caja = max(max_caja, int(caja))
            
            return max_caja + 1
        except Exception as e:
            print(f"Error al obtener número de caja: {e}")
            return 1  # Retorna 1 en caso de error (comenzar desde 1)