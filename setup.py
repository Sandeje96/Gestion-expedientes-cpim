#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script de configuración para el Sistema de Gestión CPIM
Este script crea la estructura de carpetas y el archivo Excel inicial
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font
from pathlib import Path

def main():
    print("Iniciando configuración del Sistema de Gestión CPIM...")
    
    # Obtener la ruta base del proyecto
    base_path = Path(__file__).parent
    print(f"Ruta base del proyecto: {base_path}")
    
    # Crear carpetas principales
    folders = ["data", "templates", "trabajos", "trabajos/OBRA NUEVA", "trabajos/REGISTRACION"]
    for folder in folders:
        folder_path = base_path / folder
        folder_path.mkdir(exist_ok=True)
        print(f"Carpeta creada/verificada: {folder_path}")
    
    # Crear archivo Excel
    excel_path = base_path / "data" / "registros.xlsx"
    
    # Si existe, intentar eliminarlo primero
    if excel_path.exists():
        try:
            os.remove(excel_path)
            print(f"Archivo Excel existente eliminado: {excel_path}")
        except Exception as e:
            print(f"No se pudo eliminar el archivo Excel existente: {e}")
            # Crear con otro nombre
            excel_path = base_path / "data" / "registros_new.xlsx"
    
    try:
        # Crear un nuevo workbook
        workbook = openpyxl.Workbook()
        
        # Configurar la hoja de Obras en general
        obras_sheet = workbook.active
        obras_sheet.title = "Obras en general"
        
        # Crear hoja para informes técnicos
        informes_sheet = workbook.create_sheet("Informes técnicos")
        
        # Configurar encabezados para Obras en general
        headers_obras = [
            "Fecha", "Profesión", "Formato", "Nro de Copias", "Tipo de trabajo", 
            "Nombre del Profesional", "Nombre del Comitente", "Ubicación", 
            "Nro de expte municipal", "Nro de sistema GOP", "Nro de partida inmobiliaria",
            "Tasa de sellado", "Tasa de visado", 
            "Visado de instalacion de Gas", "Visado de instalacion de Salubridad", 
            "Visado de instalacion electrica", "Visado de instalacion electromecanica",
            "Estado pago sellado", "Estado pago visado",
            "Nro de expediente CPIM", "Fecha de salida", "Persona que retira", 
            "Nro de Caja", "Ruta de carpeta"
        ]
        
        # Configurar encabezados para Informes técnicos
        headers_informes = [
            "Fecha", "Profesión", "Formato", "Nro de Copias", "Tipo de trabajo", 
            "Detalle", "Profesional", "Comitente", "Tasa de sellado", 
            "Estado de pago", "Nro de expediente CPIM", "Fecha de salida", 
            "Persona que retira", "Nro de Caja", "Ruta de carpeta"
        ]
        
        # Aplicar encabezados a las hojas
        for idx, header in enumerate(headers_obras, 1):
            cell = obras_sheet.cell(row=1, column=idx, value=header)
            cell.font = Font(bold=True)
        
        for idx, header in enumerate(headers_informes, 1):
            cell = informes_sheet.cell(row=1, column=idx, value=header)
            cell.font = Font(bold=True)
        
        # Guardar el archivo
        workbook.save(excel_path)
        print(f"Archivo Excel creado correctamente: {excel_path}")
        
        # Verificar que se puede abrir
        test_workbook = openpyxl.load_workbook(excel_path)
        test_workbook.close()
        print("Verificación exitosa: El archivo Excel se puede abrir correctamente.")
        
    except Exception as e:
        print(f"Error al crear el archivo Excel: {e}")
        return False
    
    print("Configuración completada con éxito.")
    print("Ahora puedes ejecutar 'python main.py' para iniciar el programa.")
    return True

if __name__ == "__main__":
    success = main()
    if not success:
        print("La configuración falló. Por favor revisa los mensajes de error.")
    input("Presiona Enter para continuar...")